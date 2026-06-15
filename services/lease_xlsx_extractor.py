import uuid
from datetime import datetime, date
from typing import Any

import openpyxl


def extract_lease_data_from_excels(rules_xlsx_path: str, data_xlsx_path: str) -> dict:
    rules = _load_rules(rules_xlsx_path)
    tables = _load_tables(data_xlsx_path)

    base_rows = tables.get('基础信息', [])
    fee_rows = tables.get('费用信息', [])

    fee_by_lease: dict[str, list[dict]] = {}
    for r in fee_rows:
        lease_no = _as_str(r.get('租约编号'))
        if not lease_no:
            continue
        fee_by_lease.setdefault(lease_no, []).append(r)

    base_rules = rules.get('基础信息', [])
    rights_rules = rules.get('权益字段提取', [])
    fee_total_rules = rules.get('费用信息-录入单位时间内付费总额', [])
    fee_unit_rules = rules.get('费用信息-单价录入', [])

    base_map = {
        '行政区': '行政区',
        '租赁保证金（元）': '租赁保证金（元）',
    }

    fee_map = {
        '费用类型': '费用类别',
        '付费开始时间': '付费开始日期',
        '付费结束时间': '付费结束日期',
        '单位时间内付费金额': '单位时间内付费金额',
        '付费频率': '付费频率',
        '付费单价': '付费单价',
        '备注': '备注',
        '合同编号': '合同编号',
        '城市': '城市',
        '楼宇名称': '楼宇名称',
        '当前租约起租日': '当前租约起租日',
        '合同约定到期日': '合同约定到期日',
        '收款方': '收款方',
        '汇率': '汇率',
        '税率（%）': '税率（%）',
        '税费': '税费',
        '不含税金额': '不含税金额',
    }

    leases: list[dict] = []
    seen: set[str] = set()

    for row in base_rows:
        lease_no = _as_str(row.get('租约编号'))
        if not lease_no:
            continue
        seen.add(lease_no)
        lease_id = uuid.uuid4().hex

        base_data = _extract_fields_from_row(row, base_rules, base_map, lease_id)
        rights_data = _extract_fields_from_row(row, rights_rules, base_map, lease_id, include_id=False)

        fees = []
        for fr in fee_by_lease.get(lease_no, []):
            input_mode = _as_str(fr.get('录入方式'))
            use_unit = bool(input_mode and ('单价' in input_mode))
            fee_rules = fee_unit_rules if use_unit else fee_total_rules
            fee_data = _extract_fields_from_row(fr, fee_rules, fee_map, lease_id)
            if fee_data:
                fees.append(fee_data)

        leases.append(
            {
                'lease_no': lease_no,
                'id': lease_id,
                'base': base_data,
                'rights': rights_data,
                'fees': fees,
            }
        )

    for lease_no, frs in fee_by_lease.items():
        if lease_no in seen:
            continue
        lease_id = uuid.uuid4().hex
        fees = []
        for fr in frs:
            input_mode = _as_str(fr.get('录入方式'))
            use_unit = bool(input_mode and ('单价' in input_mode))
            fee_rules = fee_unit_rules if use_unit else fee_total_rules
            fee_data = _extract_fields_from_row(fr, fee_rules, fee_map, lease_id)
            if fee_data:
                fees.append(fee_data)
        leases.append({'lease_no': lease_no, 'id': lease_id, 'base': {}, 'rights': {}, 'fees': fees})

    return {
        'success': True,
        'leases': leases,
        'meta': {
            'lease_count': len(leases),
            'fee_count': sum(len(l.get('fees') or []) for l in leases),
            'generated_at': datetime.utcnow().isoformat()
        }
    }


def _load_rules(xlsx_path: str) -> dict[str, list[str]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: dict[str, list[str]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if not rows or not rows[0]:
            continue
        headers = [_as_str(h) for h in rows[0]]
        if not headers:
            continue
        try:
            idx_name = headers.index('字段名称')
        except ValueError:
            continue

        fields: list[str] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) <= idx_name:
                continue
            field = _as_str(r[idx_name])
            if not field:
                continue
            fields.append(field)
        out[name] = fields
    return out


def _load_tables(xlsx_path: str) -> dict[str, list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: dict[str, list[dict[str, Any]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            continue
        headers = [_as_str(h) for h in header_row]
        if not any(headers):
            continue
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            if all([c is None or _as_str(c) == '' for c in r]):
                continue
            item: dict[str, Any] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = r[i] if i < len(r) else None
                item[h] = _normalize_cell(v)
            rows.append(item)
        out[name] = rows
    return out


def _extract_fields_from_row(row: dict, fields: list[str], mapping: dict[str, str], lease_id: str, include_id: bool = True) -> dict:
    if not fields:
        return {}
    out: dict[str, Any] = {}
    for f in fields:
        if f == 'ID':
            if include_id:
                out[f] = lease_id
            continue
        col = mapping.get(f, f)
        v = row.get(col)
        if v is None and col != f:
            v = row.get(f)
        out[f] = v
    return out


def _normalize_cell(v: Any) -> Any:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def _as_str(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()
