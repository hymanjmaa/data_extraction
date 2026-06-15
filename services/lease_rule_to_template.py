import openpyxl
from typing import Any


def lease_rules_to_template_fields(rules_xlsx_path: str) -> list[dict]:
    """
    从租约规则Excel生成模板字段列表
    返回格式: [{"name": "...", "description": "...", "type": "...", "required": False}, ...]
    """
    wb = openpyxl.load_workbook(rules_xlsx_path, data_only=True)
    fields = []
    seen: set[str] = set()

    sheet_map = {
        '基础信息': 'base',
        '费用信息-录入单位时间内付费总额': 'fee_total',
        '费用信息-单价录入': 'fee_unit',
        '权益字段提取': 'rights',
        '事件导入': 'events',
    }

    for sheet_name, sheet_type in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            continue

        headers = rows[0]
        if len(headers) < 3:
            continue

        name_col_idx = 0
        type_col_idx = 1
        rule_col_idx = 2

        for row in rows[1:]:
            if len(row) <= rule_col_idx:
                continue
            field_name = _as_str(row[name_col_idx])
            field_type = _as_str(row[type_col_idx])
            rule_desc = _as_str(row[rule_col_idx])

            if not field_name or field_name in seen:
                continue
            seen.add(field_name)

            is_required = _is_required(field_name, sheet_type)
            fields.append({
                'name': field_name,
                'description': f"[{sheet_name}] {rule_desc}" if rule_desc else f"[{sheet_name}]",
                'type': _normalize_field_type(field_type),
                'required': is_required,
                'validation_rules': [],
            })

    return fields


def _as_str(v: Any) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _is_required(field_name: str, sheet_type: str) -> bool:
    required_base = {'租约编号', '承租方', '城市', '楼宇名称', '计租面积', '合同约定到期日', '当前租约起租日'}
    required_fee = {'费用类型', '付费开始时间', '付费结束时间'}
    required_rights = {'是否需要还原', '有无提前解约权', '有无优先续租权'}

    if sheet_type == 'base':
        return field_name in required_base
    if sheet_type in ('fee_total', 'fee_unit'):
        return field_name in required_fee
    if sheet_type == 'rights':
        return field_name in required_rights
    return False


def _normalize_field_type(raw: str) -> str:
    t = raw.lower().strip()
    if 'int' in t:
        return 'integer'
    if 'decimal' in t or 'float' in t or '金额' in t or '面积' in t or '率' in t:
        return 'number'
    if 'bool' in t:
        return 'boolean'
    if 'date' in t:
        return 'date'
    return 'string'
