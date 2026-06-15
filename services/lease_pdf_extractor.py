import os
import uuid
from datetime import datetime, date
from typing import Any
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl

from services.extraction_engine import AIExtractionEngine
from utils.document_parser import DocumentParser


def extract_lease_from_pdfs(
    rule_xlsx_path: str,
    pdf_paths: list[str],
    output_template_xlsx_path: str,
    output_dir: str = "uploads",
) -> dict:
    if not os.path.exists(rule_xlsx_path):
        return {'success': False, 'error': '规则文件不存在'}
    if not os.path.exists(output_template_xlsx_path):
        return {'success': False, 'error': '输出模板文件不存在'}
    if not pdf_paths:
        return {'success': False, 'error': '未提供PDF文件'}

    engine = AIExtractionEngine()
    if not engine.api_key:
        return {'success': False, 'error': 'LLM API密钥未配置'}

    rule_text = _load_rules_as_text(rule_xlsx_path, max_chars=25000)
    wb_template = openpyxl.load_workbook(output_template_xlsx_path)

    if '基础信息' not in wb_template.sheetnames or '费用信息' not in wb_template.sheetnames:
        return {'success': False, 'error': '输出模板缺少sheet：基础信息/费用信息'}

    base_ws = wb_template['基础信息']
    fee_ws = wb_template['费用信息']
    base_headers = _read_header_row(base_ws)
    fee_headers = _read_header_row(fee_ws)

    _clear_rows(base_ws, start_row=2)
    _clear_rows(fee_ws, start_row=2)

    lease_count = 0
    fee_count = 0
    errors = []

    def _process_one_pdf(pdf_path: str):
        parsed = DocumentParser.auto_parse(pdf_path)
        if not parsed.get('success'):
            return {'pdf': pdf_path, 'stage': 'parse', 'error': parsed.get('error')}
        pdf_text = parsed.get('text', '')
        if not pdf_text:
            return {'pdf': pdf_path, 'stage': 'parse', 'error': 'empty_text'}

        extracted = _extract_one_lease(engine, pdf_text, rule_text, base_headers, fee_headers, os.path.basename(pdf_path))
        if not extracted.get('success'):
            return {'pdf': pdf_path, 'stage': 'extract', 'error': extracted.get('error')}

        lease = extracted.get('lease') or {}
        fees = extracted.get('fees') or []

        lease_row = {h: lease.get(h, '') for h in base_headers}
        if lease_row.get('备注') in [None, '']:
            lease_row['备注'] = lease.get('提取备注', '') or ''

        if not _has_meaningful_values(lease_row):
            return {'pdf': pdf_path, 'stage': 'extract', 'error': 'empty_fields'}

        return {
            'pdf': pdf_path,
            'success': True,
            'lease_row': lease_row,
            'fees': [{h: fee.get(h, '') for h in fee_headers} for fee in fees],
        }

    max_workers = max(1, min(len(pdf_paths), 4))
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = {pool.submit(_process_one_pdf, p): p for p in pdf_paths}
        for fut in as_completed(futures):
            result = fut.result()
            if not result.get('success'):
                errors.append(result)
                continue
            _append_row(base_ws, base_headers, result['lease_row'])
            lease_count += 1
            for fee_row in result['fees']:
                _append_row(fee_ws, fee_headers, fee_row)
                fee_count += 1

    os.makedirs(output_dir, exist_ok=True)
    out_name = f"租约提取结果_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.xlsx"
    out_path = os.path.join(output_dir, out_name)
    wb_template.save(out_path)

    if lease_count == 0:
        return {
            'success': False,
            'error': '未能从PDF提取到任何租约数据',
            'details': errors[:50],
            'output_file': out_name
        }

    return {
        'success': True,
        'output_file': out_name,
        'meta': {
            'lease_count': lease_count,
            'fee_count': fee_count,
            'generated_at': datetime.utcnow().isoformat()
        }
    }


def _extract_one_lease(
    engine: AIExtractionEngine,
    pdf_text: str,
    rule_text: str,
    base_headers: list[str],
    fee_headers: list[str],
    file_name: str,
) -> dict:
    if pdf_text and len(pdf_text) > 50000:
        pdf_text = f"{pdf_text[:30000]}\n...\n{pdf_text[-5000:]}"

    lease: dict[str, Any] = {}
    chunks = _chunk_list(base_headers, 18)
    for idx, fields in enumerate(chunks):
        part = _extract_base_fields(engine, pdf_text, rule_text, fields, file_name, idx)
        if part is None:
            continue
        lease.update(part)

    fees = _extract_fees(engine, pdf_text, rule_text, fee_headers, file_name)

    lease = {str(k): _normalize_cell(v) for k, v in lease.items()}
    normalized_fees = []
    for fee in fees:
        if not isinstance(fee, dict):
            continue
        normalized_fees.append({str(k): _normalize_cell(v) for k, v in fee.items()})

    if not lease:
        return {'success': False, 'error': 'LLM未返回基础信息字段'}

    return {'success': True, 'lease': lease, 'fees': normalized_fees}


def _extract_base_fields(engine: AIExtractionEngine, pdf_text: str, rule_text: str, fields: list[str], file_name: str, chunk_index: int) -> dict | None:
    field_lines = "\n".join([f"- {h}" for h in fields])
    prompt = f"""你是一个合同/租约信息抽取助手。你只输出JSON，不要输出其它文本。

请根据“租约字段提取规则”与PDF文本，抽取下面这些字段。
字段名必须严格使用给定字段名，缺失用空字符串。

规则：
{rule_text}

PDF文件名：
{file_name}

PDF文本：
{pdf_text}

字段列表：
{field_lines}

输出JSON格式：
{{"data": {{"<字段名>": "<值或空字符串>"}}}}
"""

    res = engine.run_json_prompt(prompt, temperature=0.1, max_tokens=1800)
    # #region debug-point B:llm-base-chunk
    import json as _j, urllib.request as _ur; _p='.dbg/lease-pdf-500.env'; _u,_s='http://127.0.0.1:7777/event','lease-pdf-500'; exec("try:\n with open(_p) as f: c=f.read(); _u=next((l.split('=',1)[1] for l in c.split('\\n') if l.startswith('DEBUG_SERVER_URL=')),_u); _s=next((l.split('=',1)[1] for l in c.split('\\n') if l.startswith('DEBUG_SESSION_ID=')),_s)\nexcept: pass"); _ur.urlopen(_ur.Request(_u, data=_j.dumps({'sessionId':_s,'runId':'post','hypothesisId':'B','location':'lease_pdf_extractor.py:_extract_base_fields','msg':'[DEBUG] base fields chunk','data':{'chunk':chunk_index,'field_count':len(fields),'success':bool(res.get('success')),'error':res.get('error')},'ts':int(__import__('time').time()*1000)}).encode(), headers={'Content-Type':'application/json'})).read()
    # #endregion

    if not res.get('success') or not isinstance(res.get('data'), dict):
        return None
    data = res.get('data') or {}
    part = data.get('data') if isinstance(data.get('data'), dict) else None
    if not isinstance(part, dict):
        return None
    return part


def _extract_fees(engine: AIExtractionEngine, pdf_text: str, rule_text: str, fee_headers: list[str], file_name: str) -> list[dict]:
    fee_fields = "\n".join([f"- {h}" for h in fee_headers])
    prompt = f"""你是一个合同/租约信息抽取助手。你只输出JSON，不要输出其它文本。

请从PDF文本中抽取“费用信息”明细，可能有多条。
字段名必须严格使用给定字段名，缺失用空字符串；如果合同没有费用明细则输出空数组。

规则：
{rule_text}

PDF文件名：
{file_name}

PDF文本：
{pdf_text}

费用字段：
{fee_fields}

输出JSON格式：
{{"fees":[{{"<费用字段>":"<值或空字符串>"}}]}}
"""
    res = engine.run_json_prompt(prompt, temperature=0.1, max_tokens=2600)
    # #region debug-point B:llm-fees
    import json as _j, urllib.request as _ur; _p='.dbg/lease-pdf-500.env'; _u,_s='http://127.0.0.1:7777/event','lease-pdf-500'; exec("try:\n with open(_p) as f: c=f.read(); _u=next((l.split('=',1)[1] for l in c.split('\\n') if l.startswith('DEBUG_SERVER_URL=')),_u); _s=next((l.split('=',1)[1] for l in c.split('\\n') if l.startswith('DEBUG_SESSION_ID=')),_s)\nexcept: pass"); _ur.urlopen(_ur.Request(_u, data=_j.dumps({'sessionId':_s,'runId':'post','hypothesisId':'B','location':'lease_pdf_extractor.py:_extract_fees','msg':'[DEBUG] fees extract','data':{'success':bool(res.get('success')),'error':res.get('error')},'ts':int(__import__('time').time()*1000)}).encode(), headers={'Content-Type':'application/json'})).read()
    # #endregion

    if not res.get('success') or not isinstance(res.get('data'), dict):
        return []
    data = res.get('data') or {}
    fees = data.get('fees') if isinstance(data.get('fees'), list) else []
    out = []
    for fee in fees:
        if isinstance(fee, dict):
            out.append(fee)
    return out


def _chunk_list(items: list[str], size: int) -> list[list[str]]:
    if size <= 0:
        return [items]
    return [items[i:i + size] for i in range(0, len(items), size)]


def _has_meaningful_values(row_dict: dict) -> bool:
    keys = ['租约编号', '合同编号', '承租方', '楼宇名称', '详细地址', '当前租约起租日', '合同约定到期日']
    for k in keys:
        v = row_dict.get(k)
        if v not in [None, '', '-', 'N/A', 'n/a']:
            return True
    for v in row_dict.values():
        if v not in [None, '', '-', 'N/A', 'n/a']:
            return True
    return False


def _load_rules_as_text(xlsx_path: str, max_chars: int = 20000) -> str:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    chunks: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            continue
        headers = [_as_str(h) for h in header]
        if '字段名称' not in headers:
            continue
        idx_name = headers.index('字段名称')
        idx_rule = None
        for cand in ['字段抽取规则', '字段提取规则']:
            if cand in headers:
                idx_rule = headers.index(cand)
                break
        if idx_rule is None:
            continue

        lines: list[str] = [f"【{name}】"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or len(r) <= idx_name:
                continue
            fname = _as_str(r[idx_name])
            if not fname:
                continue
            rule = _as_str(r[idx_rule]) if idx_rule < len(r) else ''
            if rule:
                lines.append(f"- {fname}: {rule}")
            else:
                lines.append(f"- {fname}")
        text = "\n".join(lines)
        chunks.append(text)

    joined = "\n\n".join(chunks)
    if len(joined) > max_chars:
        joined = joined[:max_chars]
    return joined


def _read_header_row(ws) -> list[str]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return []
    return [_as_str(v) for v in header_row if _as_str(v)]


def _clear_rows(ws, start_row: int = 2):
    if ws.max_row < start_row:
        return
    ws.delete_rows(start_row, ws.max_row - start_row + 1)


def _append_row(ws, headers: list[str], row_dict: dict):
    row = []
    for h in headers:
        v = row_dict.get(h, '')
        row.append(_normalize_cell(v))
    ws.append(row)


def _normalize_cell(v: Any) -> Any:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return v


def _as_str(v: Any) -> str:
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()
